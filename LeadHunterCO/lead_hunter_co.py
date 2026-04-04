#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════╗
║  LEAD HUNTER CO v1.0 — Google Maps Lead Generation Tool     ║
║  Para: BlackCats Agency / Skuall Studio                     ║
║  Créditos: Cristian Saenz                                   ║
╚══════════════════════════════════════════════════════════════╝

Scraper automatizado de Google Maps que identifica negocios
con alta probabilidad de necesitar servicios web.

Requiere: Google Places API Key (https://console.cloud.google.com)
"""

import requests
import json
import time
import os
import sys
import re
import logging
from datetime import datetime
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

# ============================================================
# CONFIGURACIÓN
# ============================================================

# Ciudades colombianas a escanear
CIUDADES = {
    "Bogotá": {"lat": 4.6097, "lng": -74.0817},
    "Medellín": {"lat": 6.2442, "lng": -75.5812},
    "Cali": {"lat": 3.4516, "lng": -76.5320},
    "Barranquilla": {"lat": 10.9685, "lng": -74.7813},
    "Cartagena": {"lat": 10.3910, "lng": -75.5144},
    "Bucaramanga": {"lat": 7.1254, "lng": -73.1198},
    "Santa Marta": {"lat": 11.2404, "lng": -74.1990},
    "Pereira": {"lat": 4.8087, "lng": -75.6906},
    "Manizales": {"lat": 5.0689, "lng": -75.5174},
    "Ibagué": {"lat": 4.4389, "lng": -75.2322},
    "Villavicencio": {"lat": 4.1420, "lng": -73.6266},
    "Cúcuta": {"lat": 7.8939, "lng": -72.5078},
}

# Nichos y queries de búsqueda
NICHOS = {
    "Plomería": ["plomero", "plomería", "destapes"],
    "HVAC": ["aire acondicionado", "refrigeración", "HVAC"],
    "Electricista": ["electricista", "eléctricos", "instalaciones eléctricas"],
    "Medicina Estética": ["medicina estética", "centro estético", "dermatología estética"],
    "Spa / Bienestar": ["spa", "masajes", "wellness"],
    "Odontología": ["odontología", "dentista", "ortodoncia"],
    "Veterinaria": ["veterinaria", "clínica veterinaria"],
    "Gastronomía": ["restaurante", "asadero", "marisquería"],
    "Fitness": ["gimnasio", "crossfit", "pilates"],
    "Barbería": ["barbería", "peluquería"],
    "Fisioterapia": ["fisioterapia", "rehabilitación física"],
    "Taller Automotriz": ["taller mecánico", "mecánica automotriz"],
    "Lavado Autos": ["lavado de autos", "car wash", "detailing"],
    "Hotelería": ["hotel boutique", "hostal", "hospedaje"],
}

# Scoring weights
SCORING = {
    "sin_web": 25,         # No tiene sitio web
    "volumen_reseñas": 25, # Cantidad de reseñas
    "rating_alto": 10,     # Rating >= 4.3
    "nicho_alto_impacto": 15,  # Nicho donde web impacta ingresos
    "dolor_en_reseñas": 10,    # Señales de dolor en reseñas
    "mercado_internacional": 5, # Reseñas en inglés
    "ticket_potencial": 5,     # Ticket estimado alto
    "urgencia": 5,             # Señales de urgencia
}

NICHOS_ALTO_IMPACTO = [
    "Medicina Estética", "Spa / Bienestar", "Odontología",
    "Gastronomía", "Veterinaria", "Hotelería"
]

# Palabras de dolor en reseñas
DOLOR_KEYWORDS = [
    "espera", "lento", "confus", "difícil", "reservar", "reserva",
    "comunicación", "cobr", "caro", "engaño", "estaf", "malo",
    "terrible", "horrible", "wait", "slow", "confus", "difficult",
    "booking", "expensive", "scam", "overpriced", "dirty",
    "appointment", "schedule", "contact", "website", "online",
    "no answer", "no responde", "no contestan",
]

# ============================================================
# CLASES PRINCIPALES
# ============================================================

class GooglePlacesClient:
    """Cliente para Google Places API (New)"""

    BASE_URL = "https://maps.googleapis.com/maps/api/place"

    def __init__(self, api_key: str):
        self.api_key = api_key
        self.session = requests.Session()
        self.request_count = 0

    def text_search(self, query: str, location: dict, radius: int = 30000) -> list:
        """Busca negocios por texto + ubicación"""
        url = f"{self.BASE_URL}/textsearch/json"
        params = {
            "query": query,
            "location": f"{location['lat']},{location['lng']}",
            "radius": radius,
            "key": self.api_key,
            "language": "es",
        }
        results = []
        try:
            resp = self.session.get(url, params=params, timeout=15)
            self.request_count += 1
            data = resp.json()
            if data.get("status") == "OK":
                results = data.get("results", [])
            elif data.get("status") == "REQUEST_DENIED":
                logging.error(f"API Key inválida o sin permisos: {data.get('error_message')}")
            time.sleep(0.3)  # Rate limiting
        except Exception as e:
            logging.error(f"Error en text_search: {e}")
        return results

    def get_details(self, place_id: str) -> dict:
        """Obtiene detalles completos de un negocio"""
        url = f"{self.BASE_URL}/details/json"
        params = {
            "place_id": place_id,
            "fields": "name,formatted_address,formatted_phone_number,website,"
                      "rating,user_ratings_total,reviews,opening_hours,types,"
                      "business_status,url,geometry",
            "key": self.api_key,
            "language": "es",
            "reviews_sort": "newest",
        }
        try:
            resp = self.session.get(url, params=params, timeout=15)
            self.request_count += 1
            data = resp.json()
            if data.get("status") == "OK":
                time.sleep(0.2)
                return data.get("result", {})
        except Exception as e:
            logging.error(f"Error en get_details: {e}")
        return {}


class WebsiteValidator:
    """Valida si un sitio web existe y su calidad básica"""

    @staticmethod
    def check_website(url: str) -> dict:
        """Verifica existencia y estado básico de un sitio web"""
        result = {
            "exists": False,
            "status_code": None,
            "is_https": False,
            "load_time_ms": None,
            "has_mobile_meta": False,
            "error": None,
        }
        if not url or url in ["—", "N/A", ""]:
            return result
        try:
            if not url.startswith("http"):
                url = f"https://{url}"
            start = time.time()
            resp = requests.get(url, timeout=10, allow_redirects=True,
                              headers={"User-Agent": "LeadHunterCO/1.0"})
            load_time = (time.time() - start) * 1000
            result["exists"] = resp.status_code < 400
            result["status_code"] = resp.status_code
            result["is_https"] = resp.url.startswith("https")
            result["load_time_ms"] = round(load_time)
            # Check mobile viewport meta
            content = resp.text[:5000].lower()
            result["has_mobile_meta"] = "viewport" in content
        except requests.exceptions.SSLError:
            result["error"] = "SSL_ERROR"
        except requests.exceptions.ConnectionError:
            result["error"] = "CONNECTION_ERROR"
        except requests.exceptions.Timeout:
            result["error"] = "TIMEOUT"
        except Exception as e:
            result["error"] = str(e)[:50]
        return result


class LeadScorer:
    """Calcula score de probabilidad de ser buen lead"""

    @staticmethod
    def calculate_score(lead: dict) -> int:
        score = 0

        # 1. Sin web (25 pts)
        if not lead.get("website"):
            score += SCORING["sin_web"]
        elif lead.get("web_check", {}).get("load_time_ms", 0) > 4000:
            score += 15  # Web lenta = oportunidad de rediseño
        elif not lead.get("web_check", {}).get("has_mobile_meta"):
            score += 12  # Sin mobile = oportunidad

        # 2. Volumen de reseñas (25 pts)
        reviews = lead.get("rating_count", 0) or 0
        if reviews >= 500:
            score += 25
        elif reviews >= 100:
            score += 20
        elif reviews >= 50:
            score += 15
        elif reviews >= 20:
            score += 10
        else:
            score += 5

        # 3. Rating alto (10 pts)
        rating = lead.get("rating", 0) or 0
        if rating >= 4.7:
            score += 10
        elif rating >= 4.3:
            score += 7
        elif rating >= 4.0:
            score += 5

        # 4. Nicho alto impacto (15 pts)
        if lead.get("nicho") in NICHOS_ALTO_IMPACTO:
            score += 15
        else:
            score += 8

        # 5. Dolor en reseñas (10 pts)
        dolor_count = lead.get("dolor_signals", 0)
        if dolor_count >= 3:
            score += 10
        elif dolor_count >= 1:
            score += 6

        # 6. Mercado internacional (5 pts)
        if lead.get("has_english_reviews"):
            score += 5

        # 7. Ticket potencial (5 pts)
        if lead.get("nicho") in ["Medicina Estética", "Odontología", "Hotelería"]:
            score += 5
        elif lead.get("nicho") in ["Gastronomía", "Spa / Bienestar"]:
            score += 3

        # 8. Urgencia (5 pts) — si cambió dirección, o competidor tiene web
        if lead.get("urgency_signal"):
            score += 5

        return min(score, 100)

    @staticmethod
    def get_priority(score: int) -> str:
        if score >= 80:
            return "🔴 ALTA"
        elif score >= 60:
            return "🟡 MEDIA"
        elif score >= 40:
            return "🟢 BAJA"
        return "⚪ DESCARTAR"

    @staticmethod
    def analyze_reviews(reviews: list) -> dict:
        """Analiza reseñas para señales de dolor y mercado internacional"""
        dolor_count = 0
        english_count = 0
        dolor_examples = []

        for review in (reviews or []):
            text = review.get("text", "") if isinstance(review, dict) else str(review)
            text_lower = text.lower()

            # Detectar dolor
            for kw in DOLOR_KEYWORDS:
                if kw in text_lower:
                    dolor_count += 1
                    if len(dolor_examples) < 3:
                        snippet = text[:120].replace("\n", " ")
                        dolor_examples.append(f'"{snippet}..."')
                    break

            # Detectar inglés (heurística simple)
            english_words = ["the", "was", "very", "good", "great", "excellent",
                           "service", "recommend", "place", "food"]
            eng_matches = sum(1 for w in english_words if w in text_lower)
            if eng_matches >= 3:
                english_count += 1

        return {
            "dolor_signals": dolor_count,
            "has_english_reviews": english_count >= 2,
            "english_review_pct": round(english_count / max(len(reviews or []), 1) * 100),
            "dolor_examples": dolor_examples,
        }


class ExcelExporter:
    """Exporta leads a Excel estilizado"""

    @staticmethod
    def export(leads: list, output_path: str):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.chart import BarChart, Reference
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("ERROR: pip install openpyxl")
            return

        wb = Workbook()
        C_NAVY = "0D1B2A"
        C_RED = "E63946"
        C_GREEN = "06D6A0"
        C_ORANGE = "F77F00"
        C_WHITE = "FFFFFF"
        C_LIGHT = "F8F9FA"
        C_GRAY = "CED4DA"

        thin = Border(
            left=Side("thin", color=C_GRAY), right=Side("thin", color=C_GRAY),
            top=Side("thin", color=C_GRAY), bottom=Side("thin", color=C_GRAY)
        )

        # ---- SHEET: DASHBOARD ----
        ws = wb.active
        ws.title = "DASHBOARD"
        ws.sheet_properties.tabColor = C_NAVY
        ws.sheet_view.showGridLines = False

        total = len(leads)
        alta = sum(1 for l in leads if "ALTA" in l.get("priority", ""))
        media = sum(1 for l in leads if "MEDIA" in l.get("priority", ""))
        sin_web = sum(1 for l in leads if not l.get("website"))
        total_reviews = sum(l.get("rating_count", 0) or 0 for l in leads)
        ciudades = len(set(l.get("city", "") for l in leads))

        ws.merge_cells("A1:L1")
        ws["A1"] = f"📊 LEAD HUNTER CO — REPORTE {datetime.now().strftime('%d/%m/%Y')}"
        ws["A1"].font = Font("Arial", 20, bold=True, color=C_WHITE)
        ws["A1"].fill = PatternFill("solid", fgColor=C_NAVY)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 50

        kpis = [
            ("TOTAL LEADS", total, C_NAVY),
            ("🔴 ALTA", alta, C_RED),
            ("🟡 MEDIA", media, C_ORANGE),
            ("SIN WEB", f"{sin_web}/{total}", C_RED),
            ("RESEÑAS", f"{total_reviews:,}", "7B2CBF"),
            ("CIUDADES", ciudades, "00B4D8"),
        ]
        for i, (label, val, color) in enumerate(kpis):
            col = i * 2 + 1
            ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
            ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)
            c1 = ws.cell(row=3, column=col, value=label)
            c1.font = Font("Arial", 9, bold=True, color=C_WHITE)
            c1.fill = PatternFill("solid", fgColor=color)
            c1.alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=col+1).fill = PatternFill("solid", fgColor=color)
            c2 = ws.cell(row=4, column=col, value=val)
            c2.font = Font("Arial", 18, bold=True, color=color)
            c2.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 25
        ws.row_dimensions[4].height = 40

        # ---- SHEET: LEADS ----
        ws2 = wb.create_sheet("Leads Completos")
        ws2.sheet_properties.tabColor = C_GREEN

        headers = [
            "ID", "Negocio", "Ciudad", "Dirección", "Nicho", "Reseñas",
            "Rating", "¿Web?", "URL Web", "Teléfono", "Horario",
            "Dolor Principal", "Reseña Clave", "¿Reseñas EN?", "% Inglés",
            "Score", "Prioridad", "Argumento Venta",
            "Web Status", "Web HTTPS", "Web Speed (ms)", "Web Mobile",
        ]
        for c, h in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=c, value=h)
            cell.font = Font("Arial", 10, bold=True, color=C_WHITE)
            cell.fill = PatternFill("solid", fgColor=C_NAVY)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin

        for r, lead in enumerate(leads, 2):
            wc = lead.get("web_check", {})
            row_data = [
                r - 1,
                lead.get("name", ""),
                lead.get("city", ""),
                lead.get("address", ""),
                lead.get("nicho", ""),
                lead.get("rating_count", 0),
                lead.get("rating", 0),
                "SÍ" if lead.get("website") else "NO",
                lead.get("website", "—"),
                lead.get("phone", ""),
                lead.get("hours_summary", ""),
                "; ".join(lead.get("dolor_examples", [])[:2]) or "Sin señales",
                lead.get("sample_review", "")[:150],
                "SÍ" if lead.get("has_english_reviews") else "NO",
                f"{lead.get('english_review_pct', 0)}%",
                lead.get("score", 0),
                lead.get("priority", ""),
                lead.get("sales_argument", ""),
                wc.get("status_code", "N/A"),
                "SÍ" if wc.get("is_https") else "NO",
                wc.get("load_time_ms", "N/A"),
                "SÍ" if wc.get("has_mobile_meta") else "NO",
            ]
            pri = lead.get("priority", "")
            for c, val in enumerate(row_data, 1):
                cell = ws2.cell(row=r, column=c, value=val)
                cell.font = Font("Arial", 9, color="1A1A2E")
                cell.alignment = Alignment(
                    horizontal="left" if c in [2,4,12,13,18] else "center",
                    vertical="center", wrap_text=True
                )
                cell.border = thin
                if "ALTA" in pri:
                    cell.fill = PatternFill("solid", fgColor="F8D7DA")
                elif "MEDIA" in pri:
                    cell.fill = PatternFill("solid", fgColor="FFF3CD")
                elif r % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=C_LIGHT)

            if "ALTA" in pri:
                ws2.cell(row=r, column=17).font = Font("Arial", 9, bold=True, color=C_RED)
                ws2.cell(row=r, column=16).font = Font("Arial", 10, bold=True, color=C_RED)

        widths = [4,30,14,35,18,8,7,6,25,18,20,45,50,8,8,7,12,55,8,6,10,6]
        for i, w in enumerate(widths, 1):
            ws2.column_dimensions[get_column_letter(i)].width = w
        ws2.freeze_panes = "C2"
        ws2.auto_filter.ref = f"A1:V{len(leads)+1}"
        for r in range(2, len(leads) + 2):
            ws2.row_dimensions[r].height = 65

        # ---- SHEET: CRM PIPELINE ----
        ws3 = wb.create_sheet("Pipeline CRM")
        ws3.sheet_properties.tabColor = "00B4D8"
        crm_h = ["ID","Negocio","Ciudad","Prioridad","Score","Teléfono",
                  "Fecha Contacto","Canal","Respuesta","Follow-up 1","Follow-up 2",
                  "Estado","Ticket USD","Notas"]
        for c, h in enumerate(crm_h, 1):
            cell = ws3.cell(row=1, column=c, value=h)
            cell.font = Font("Arial", 10, bold=True, color=C_WHITE)
            cell.fill = PatternFill("solid", fgColor="00B4D8")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin
        for r, lead in enumerate(leads, 2):
            crm_row = [r-1, lead.get("name",""), lead.get("city",""),
                       lead.get("priority",""), lead.get("score",0),
                       lead.get("phone",""), "", "WhatsApp", "",
                       "", "", "Nuevo", "", ""]
            for c, val in enumerate(crm_row, 1):
                cell = ws3.cell(row=r, column=c, value=val)
                cell.font = Font("Arial", 9)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin
        crm_w = [4,30,14,12,6,18,14,12,10,14,14,10,12,30]
        for i, w in enumerate(crm_w, 1):
            ws3.column_dimensions[get_column_letter(i)].width = w
        ws3.freeze_panes = "C2"

        wb.save(output_path)
        print(f"✅ Excel guardado: {output_path}")


class SalesArgumentGenerator:
    """Genera argumentos de venta personalizados"""

    @staticmethod
    def generate(lead: dict) -> str:
        name = lead.get("name", "el negocio")
        reviews = lead.get("rating_count", 0) or 0
        rating = lead.get("rating", 0) or 0
        nicho = lead.get("nicho", "")
        has_web = bool(lead.get("website"))
        has_eng = lead.get("has_english_reviews", False)

        if not has_web and reviews >= 500:
            return (f"{reviews:,} reseñas y {rating}★ pero sin web. Cada búsqueda "
                    f"orgánica que podría llegar a {name} se va a la competencia. "
                    f"Una landing optimizada con SEO local se pagaría sola en 30 días.")
        elif not has_web and has_eng:
            return (f"Clientes internacionales ya te buscan ({reviews} reseñas, "
                    f"varias en inglés) pero sin web bilingüe no pueden reservar "
                    f"ni cotizar. Cada turista que googlea tu servicio se va a "
                    f"quien sí tiene web en inglés.")
        elif not has_web and nicho in NICHOS_ALTO_IMPACTO:
            return (f"En {nicho}, los clientes verifican online antes de ir. "
                    f"Sin web, tus {reviews} reseñas trabajan al 50%. Una web "
                    f"con galería, precios y reservas duplicaría tu conversión.")
        elif not has_web:
            return (f"{reviews} reseñas y {rating}★ sin web. Google prioriza "
                    f"negocios con sitio web en el Local Pack. Sin web, tu "
                    f"posición en Maps baja cada mes frente a competidores que sí la tienen.")
        else:
            wc = lead.get("web_check", {})
            issues = []
            if wc.get("load_time_ms", 0) > 3000:
                issues.append("velocidad móvil crítica")
            if not wc.get("has_mobile_meta"):
                issues.append("no es mobile-first")
            if not wc.get("is_https"):
                issues.append("sin HTTPS")
            if issues:
                return (f"Tu web existe pero tiene problemas: {', '.join(issues)}. "
                        f"Google te penaliza en rankings. Un upgrade técnico "
                        f"mejoraría tu posición inmediatamente.")
            return f"Web funcional pero verificar Schema, AEO y Core Web Vitals."


# ============================================================
# MOTOR PRINCIPAL
# ============================================================

class LeadHunterEngine:
    """Motor principal de búsqueda y scoring"""

    def __init__(self, api_key: str, output_dir: str = None):
        self.client = GooglePlacesClient(api_key)
        self.validator = WebsiteValidator()
        self.scorer = LeadScorer()
        self.arg_gen = SalesArgumentGenerator()
        self.output_dir = output_dir or str(Path.home() / "Documents" / "LeadHunterCO")
        os.makedirs(self.output_dir, exist_ok=True)
        self.leads = []

    def scan(self, ciudades: dict = None, nichos: dict = None,
             max_per_query: int = 10, validate_webs: bool = True,
             progress_callback=None):
        """Ejecuta el escaneo completo"""
        ciudades = ciudades or CIUDADES
        nichos = nichos or NICHOS
        total_queries = sum(len(queries) for queries in nichos.values()) * len(ciudades)
        current = 0

        print(f"\n🔍 Iniciando escaneo: {len(ciudades)} ciudades × {len(nichos)} nichos")
        print(f"   Total queries estimadas: {total_queries}")
        print(f"   Max resultados por query: {max_per_query}\n")

        seen_ids = set()

        for city_name, coords in ciudades.items():
            for nicho_name, queries in nichos.items():
                for query in queries:
                    current += 1
                    search_query = f"{query} {city_name}"
                    if progress_callback:
                        progress_callback(current, total_queries, search_query)
                    else:
                        pct = round(current / total_queries * 100)
                        print(f"  [{pct:3d}%] Buscando: {search_query}")

                    results = self.client.text_search(search_query, coords)

                    for place in results[:max_per_query]:
                        pid = place.get("place_id", "")
                        if pid in seen_ids:
                            continue
                        seen_ids.add(pid)

                        # Obtener detalles
                        details = self.client.get_details(pid)
                        if not details:
                            continue

                        # Construir lead
                        lead = self._build_lead(details, city_name, nicho_name)
                        if not lead:
                            continue

                        # Validar web si existe
                        if validate_webs and lead.get("website"):
                            lead["web_check"] = self.validator.check_website(lead["website"])
                        else:
                            lead["web_check"] = {}

                        # Analizar reseñas
                        review_analysis = self.scorer.analyze_reviews(
                            details.get("reviews", [])
                        )
                        lead.update(review_analysis)

                        # Sample review
                        reviews = details.get("reviews", [])
                        if reviews:
                            lead["sample_review"] = reviews[0].get("text", "")[:200]

                        # Calcular score
                        lead["score"] = self.scorer.calculate_score(lead)
                        lead["priority"] = self.scorer.get_priority(lead["score"])

                        # Generar argumento de venta
                        lead["sales_argument"] = self.arg_gen.generate(lead)

                        self.leads.append(lead)

        # Ordenar por score descendente
        self.leads.sort(key=lambda x: x.get("score", 0), reverse=True)

        # Eliminar duplicados por nombre similar
        self.leads = self._deduplicate(self.leads)

        print(f"\n✅ Escaneo completado:")
        print(f"   Leads encontrados: {len(self.leads)}")
        print(f"   API requests: {self.client.request_count}")
        print(f"   Alta prioridad: {sum(1 for l in self.leads if 'ALTA' in l.get('priority',''))}")

        return self.leads

    def _build_lead(self, details: dict, city: str, nicho: str) -> dict:
        """Construye un lead desde los detalles de Google Places"""
        name = details.get("name", "")
        if not name:
            return None

        # Filtrar negocios cerrados
        if details.get("business_status") == "CLOSED_PERMANENTLY":
            return None

        hours = details.get("opening_hours", {})
        weekday_text = hours.get("weekday_text", [])
        hours_summary = "; ".join(weekday_text[:2]) if weekday_text else "No especificado"

        return {
            "name": name,
            "city": city,
            "address": details.get("formatted_address", ""),
            "nicho": nicho,
            "rating": details.get("rating"),
            "rating_count": details.get("user_ratings_total", 0),
            "website": details.get("website"),
            "phone": details.get("formatted_phone_number", ""),
            "hours_summary": hours_summary,
            "maps_url": details.get("url", ""),
            "place_id": details.get("place_id", ""),
            "types": details.get("types", []),
        }

    def _deduplicate(self, leads: list) -> list:
        """Elimina duplicados por nombre similar"""
        seen = set()
        unique = []
        for lead in leads:
            key = re.sub(r'[^a-z0-9]', '', lead["name"].lower())[:20]
            if key not in seen:
                seen.add(key)
                unique.append(lead)
        return unique

    def export_excel(self, filename: str = None):
        """Exporta a Excel"""
        if not self.leads:
            print("⚠️ No hay leads para exportar")
            return
        filename = filename or f"leads_colombia_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        path = os.path.join(self.output_dir, filename)
        ExcelExporter.export(self.leads, path)
        return path

    def export_json(self, filename: str = None):
        """Exporta a JSON"""
        filename = filename or f"leads_colombia_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
        path = os.path.join(self.output_dir, filename)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(self.leads, f, ensure_ascii=False, indent=2, default=str)
        print(f"✅ JSON guardado: {path}")
        return path


# ============================================================
# CLI
# ============================================================

def main():
    print("╔══════════════════════════════════════════════════════════╗")
    print("║  LEAD HUNTER CO v1.0                                    ║")
    print("║  Google Maps Lead Generation Tool                       ║")
    print("║  BlackCats Agency / Skuall Studio                       ║")
    print("╚══════════════════════════════════════════════════════════╝\n")

    # API Key
    api_key = os.environ.get("GOOGLE_PLACES_API_KEY", "")
    if not api_key:
        api_key = input("🔑 Ingresa tu Google Places API Key: ").strip()
    if not api_key:
        print("❌ Se requiere API Key. Obtén una en https://console.cloud.google.com")
        sys.exit(1)

    # Configuración
    print("\n📋 Configuración:")
    print(f"   Ciudades: {len(CIUDADES)} ({', '.join(list(CIUDADES.keys())[:5])}...)")
    print(f"   Nichos: {len(NICHOS)} ({', '.join(list(NICHOS.keys())[:5])}...)")

    custom = input("\n¿Usar configuración completa? (S/n): ").strip().lower()

    ciudades = CIUDADES
    nichos = NICHOS

    if custom == "n":
        print("\nCiudades disponibles:")
        for i, c in enumerate(CIUDADES.keys(), 1):
            print(f"  {i}. {c}")
        sel = input("Selecciona ciudades (ej: 1,2,5 o 'all'): ").strip()
        if sel != "all":
            indices = [int(x.strip()) - 1 for x in sel.split(",")]
            keys = list(CIUDADES.keys())
            ciudades = {keys[i]: CIUDADES[keys[i]] for i in indices if i < len(keys)}

        print("\nNichos disponibles:")
        for i, n in enumerate(NICHOS.keys(), 1):
            print(f"  {i}. {n}")
        sel = input("Selecciona nichos (ej: 1,3,8 o 'all'): ").strip()
        if sel != "all":
            indices = [int(x.strip()) - 1 for x in sel.split(",")]
            keys = list(NICHOS.keys())
            nichos = {keys[i]: NICHOS[keys[i]] for i in indices if i < len(keys)}

    # Ejecutar
    engine = LeadHunterEngine(api_key)
    leads = engine.scan(ciudades=ciudades, nichos=nichos)

    if leads:
        excel_path = engine.export_excel()
        json_path = engine.export_json()
        print(f"\n📁 Archivos guardados en: {engine.output_dir}")
        print(f"   Excel: {excel_path}")
        print(f"   JSON: {json_path}")
    else:
        print("\n⚠️ No se encontraron leads. Verifica tu API Key y conexión.")


if __name__ == "__main__":
    logging.basicConfig(level=logging.WARNING)
    main()
